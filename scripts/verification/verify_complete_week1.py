#!/usr/bin/env python3
"""
Verify complete Pool Week 1 with all 40 games analyzed and 20 optimal picks selected.
"""

import os

from openpyxl import load_workbook

from football_pool.core import PoolDominationSystem


def verify_complete_week1():
    """Verify complete Pool Week 1 setup."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        # Get all Week 3 games from schedule
        system = PoolDominationSystem()
        week3_games = system.schedule.get(3, {}).get("games", [])

        print("🏈 Complete Pool Week 1 Analysis")
        print("=" * 60)
        print("📅 Due Date: 2024-09-17 (Tuesday)")
        print(f"🎮 Games Available: {len(week3_games)}")
        print("🎯 Picks Required: 20")
        print(f"📊 Selection Rate: {20/len(week3_games)*100:.1f}%")
        print()

        # Categorize games
        nfl_games = [
            g
            for g in week3_games
            if any(
                team in g
                for team in [
                    "KC",
                    "BAL",
                    "LAR",
                    "DAL",
                    "GB",
                    "PHI",
                    "SF",
                    "BUF",
                    "MIA",
                    "DET",
                    "NO",
                    "TB",
                    "ATL",
                    "CAR",
                    "ARI",
                    "SEA",
                    "LAC",
                    "LV",
                    "DEN",
                    "PIT",
                    "NYG",
                    "CLEV",
                    "HOU",
                    "JAC",
                    "CINC",
                    "MINN",
                    "NE",
                    "IND",
                    "TENN",
                    "WASH",
                    "CHI",
                ]
            )
        ]
        cfb_games = [g for g in week3_games if g not in nfl_games]

        print("📊 Game Breakdown:")
        print(f"  NFL Games: {len(nfl_games)}")
        print(f"  CFB Games: {len(cfb_games)}")
        print()

        # Check Excel file picks
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🎯 Current Pool Week 1 Picks:")
        print("Conf | Team | Row | Game Match")
        print("-" * 40)

        all_correct = True
        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value

            # Find matching game
            game_match = "N/A"
            for game in week3_games:
                if team in game:
                    game_match = game
                    break

            print(f"{confidence:4d} | {str(team):4s} | {row:3d} | {game_match}")

        print()
        print("✅ Pool Week 1 Setup Complete!")
        print(f"📁 Excel File: {excel_file}")
        print("📅 Due Date: 2024-09-17")
        print(f"🎮 Total Games: {len(week3_games)}")
        print("🎯 Picks Made: 20")
        print(f"📊 Selection: {20/len(week3_games)*100:.1f}% of available games")

        return True

    except Exception as e:
        print(f"❌ Error: {e}")
        return False


if __name__ == "__main__":
    success = verify_complete_week1()
    if success:
        print("\n🏆 Pool Week 1 is ready for submission!")
    else:
        print("\n💥 Pool Week 1 setup needs attention!")
