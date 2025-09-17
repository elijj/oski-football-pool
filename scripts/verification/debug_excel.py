#!/usr/bin/env python3
"""
Debug Excel file structure to understand the correct alignment.
"""

import os

from openpyxl import load_workbook


def debug_excel_structure():
    """Debug the Excel file structure to understand correct alignment."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print(f"📊 Debugging Excel file: {excel_file}")
        print("=" * 50)

        # Check the first 25 rows and first 10 columns
        print("Excel Structure (first 25 rows, first 10 columns):")
        print("Row | Col A | Col B | Col C | Col D | Col E | Col F")
        print("-" * 50)

        for row in range(1, 26):
            values = []
            for col in range(1, 7):  # A through F
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is None:
                    values.append("")
                else:
                    values.append(str(cell_value)[:8])  # Truncate long values
            print(
                f"{row:3d} | {values[0]:6s} | {values[1]:6s} | {values[2]:6s} | {values[3]:6s} | {values[4]:6s} | {values[5]:6s}"
            )

        print("\n" + "=" * 50)

        # Check where confidence 20 should be
        print("Looking for confidence 20 placement:")
        for row in range(1, 25):
            for col in range(1, 10):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value == "KC":
                    print(f"KC found at Row {row}, Column {col}")
                if cell_value == 20:
                    print(f"Confidence 20 found at Row {row}, Column {col}")
                if cell_value == "20":
                    print(f"String '20' found at Row {row}, Column {col}")

        # Check the confidence column structure
        print("\nConfidence column (Column A) structure:")
        for row in range(1, 25):
            cell_value = ws.cell(row=row, column=1).value
            print(f"Row {row}: {cell_value}")

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")


if __name__ == "__main__":
    debug_excel_structure()
