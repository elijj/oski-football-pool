#!/usr/bin/env python3
"""
Debug the specific Excel rows that are causing issues.
"""


from openpyxl import load_workbook


def debug_excel_rows():
    """Debug the specific Excel rows."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🔍 DEBUGGING EXCEL ROWS 21-22")
        print("=" * 50)

        for row in range(21, 23):  # Rows 21-22
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value
            print(f"Row {row}: Confidence {confidence}, Team '{team}'")

        print("\n🔍 CHECKING CONTRARIAN ANALYSIS")
        import json

        with open("week_1_contrarian_analysis.json") as f:
            data = json.load(f)

        print("Last 2 picks in contrarian analysis:")
        for pick in data["optimal_picks"][-2:]:
            print(f"Team: '{pick['team']}', Confidence: {pick['confidence']}")

    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    debug_excel_rows()
