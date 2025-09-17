#!/usr/bin/env python3
"""
Verify the fixed Excel file with contrarian analysis.
"""

import os

from openpyxl import load_workbook


def verify_fixed_excel():
    """Verify the Excel file has the correct contrarian picks."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🎯 FIXED EXCEL FILE VERIFICATION")
        print("=" * 60)
        print(f"📁 File: {excel_file}")
        print("📅 Date: 2024-09-17 (Pool Week 1)")
        print("🎯 Strategy: Contrarian Analysis (FIXED)")
        print()

        print("📊 ACTUAL PICKS IN EXCEL FILE:")
        print("Conf | Team | Row | Status")
        print("-" * 40)

        # Expected contrarian picks (after team name conversion)
        expected_picks = [
            ("KC", 20),
            ("BAL", 19),
            ("LAR", 18),
            ("DAL", 17),
            ("GB", 16),
            ("SF", 15),
            ("MIA", 14),
            ("BUF", 13),
            ("DET", 12),
            ("NO", 11),
            ("TB", 10),
            ("ATL", 9),
            ("CAR", 8),
            ("ARI", 7),
            ("SEA", 6),
            ("LAC", 5),
            ("LV", 4),
            ("DEN", 3),
            ("WAS", 2),
            ("PIT", 1),
        ]

        all_correct = True

        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value

            # Find expected pick
            expected_team = None
            for exp_team, exp_conf in expected_picks:
                if exp_conf == confidence:
                    expected_team = exp_team
                    break

            status = "✅" if team == expected_team else "❌"
            if team != expected_team:
                all_correct = False

            print(f"{confidence:4d} | {str(team):4s} | {row:3d} | {status}")

        print("=" * 60)

        if all_correct:
            print("🎉 CONTRARIAN PICKS CORRECTLY PLACED!")
            print("✅ Team name conversion FIXED")
            print("✅ Contrarian analysis properly integrated")
            print("✅ Excel file ready for Pool Week 1 submission")
        else:
            print("❌ SOME PICKS STILL MISALIGNED!")
            print("Team name conversion may still have issues")

        return all_correct

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    success = verify_fixed_excel()
    if success:
        print("\n🏆 EXCEL INTEGRATION FIXED!")
    else:
        print("\n💥 EXCEL INTEGRATION STILL HAS ISSUES!")
