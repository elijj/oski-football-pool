#!/usr/bin/env python3
"""
Verify Pool Week 1 picks are correctly aligned.
"""

import os

from openpyxl import load_workbook


def verify_pool_week1():
    """Verify Pool Week 1 (NFL Week 3) picks alignment."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print(f"✅ Verifying Pool Week 1 (NFL Week 3): {excel_file}")
        print("=" * 60)

        # Expected picks from week_3_manual.json
        expected_picks = [
            ("KC", 20, 3, 2),  # KC, confidence 20, row 3, column 2 (Week 1)
            ("BAL", 19, 4, 2),  # BAL, confidence 19, row 4, column 2
            ("LAR", 18, 5, 2),  # LAR, confidence 18, row 5, column 2
            ("DAL", 17, 6, 2),  # DAL, confidence 17, row 6, column 2
            ("GB", 16, 7, 2),  # GB, confidence 16, row 7, column 2
            ("PHI", 15, 8, 2),  # PHI, confidence 15, row 8, column 2
            ("SF", 14, 9, 2),  # SF, confidence 14, row 9, column 2
            ("BUF", 13, 10, 2),  # BUF, confidence 13, row 10, column 2
            ("MIA", 12, 11, 2),  # MIA, confidence 12, row 11, column 2
            ("DET", 11, 12, 2),  # DET, confidence 11, row 12, column 2
            ("ND", 10, 13, 2),  # ND, confidence 10, row 13, column 2
            ("TB", 9, 14, 2),  # TB, confidence 9, row 14, column 2
            ("ATL", 8, 15, 2),  # ATL, confidence 8, row 15, column 2
            ("CAR", 7, 16, 2),  # CAR, confidence 7, row 16, column 2
            ("ARI", 6, 17, 2),  # ARI, confidence 6, row 17, column 2
            ("SEA", 5, 18, 2),  # SEA, confidence 5, row 18, column 2
            ("LAC", 4, 19, 2),  # LAC, confidence 4, row 19, column 2
            ("LV", 3, 20, 2),  # LV, confidence 3, row 20, column 2
            ("DEN", 2, 21, 2),  # DEN, confidence 2, row 21, column 2
            ("PIT", 1, 22, 2),  # PIT, confidence 1, row 22, column 2
        ]

        print("🎯 Checking Pool Week 1 pick alignment:")
        print("Team | Conf | Row | Col | Expected | Actual | Status")
        print("-" * 60)

        all_correct = True

        for team, confidence, expected_row, expected_col in expected_picks:
            actual_value = ws.cell(row=expected_row, column=expected_col).value
            status = "✅" if actual_value == team else "❌"

            if actual_value != team:
                all_correct = False

            print(
                f"{team:4s} | {confidence:4d} | {expected_row:3d} | {expected_col:3d} | {team:8s} | {str(actual_value):6s} | {status}"
            )

        print("=" * 60)

        if all_correct:
            print("🎉 POOL WEEK 1 PICKS CORRECTLY ALIGNED!")
            print("✅ This is Pool Week 1 (NFL Week 3 games)")
            print("✅ Due date: 2024-09-17 (Tuesday)")
            print("✅ Games: 2024-09-18 to 2024-09-22")
            print("✅ All picks align with their confidence values")
        else:
            print("❌ SOME PICKS ARE MISALIGNED!")
            print("Please check the Excel file for incorrect placements")

        return all_correct

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    success = verify_pool_week1()
    if success:
        print("\n🏆 Pool Week 1 alignment verification PASSED!")
    else:
        print("\n💥 Pool Week 1 alignment verification FAILED!")
