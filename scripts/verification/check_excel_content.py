#!/usr/bin/env python3
"""
Check the actual content of the Excel file to see what picks are there.
"""

import os

from openpyxl import load_workbook


def check_excel_content():
    """Check what's actually in the Excel file."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🔍 ACTUAL EXCEL FILE CONTENT")
        print("=" * 60)
        print(f"📁 File: {excel_file}")
        print("📅 Date: 2024-09-17")
        print()

        print("📊 ACTUAL PICKS IN EXCEL FILE:")
        print("Conf | Team | Row | Cell B | Cell A")
        print("-" * 50)

        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value
            cell_b = f"B{row}"
            cell_a = f"A{row}"

            print(f"{confidence:4d} | {str(team):4s} | {row:3d} | {cell_b:6s} | {cell_a:6s}")

        print("=" * 60)

        # Check if summary file exists
        summary_file = "Pool_Week_1_Contrarian_Analysis_Summary.md"
        if os.path.exists(summary_file):
            print(f"✅ Summary file exists: {summary_file}")
            with open(summary_file) as f:
                content = f.read()
                print(f"📄 Summary file size: {len(content)} characters")
                print("📄 First 200 characters:")
                print(content[:200] + "...")
        else:
            print(f"❌ Summary file not found: {summary_file}")

        return True

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    check_excel_content()
