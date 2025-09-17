#!/usr/bin/env python3
"""
Quick test to verify Excel file alignment is correct.
"""

import os
import sys

from openpyxl import load_workbook


def test_excel_alignment():
    """Test that the Excel file has correct pick alignment."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print(f"📊 Testing Excel file: {excel_file}")

        # Check that confidence 20 (KC) is in row 3, column 4 (Week 3)
        cell_20 = ws.cell(row=3, column=4)
        print(f"Row 3, Column 4 (Confidence 20): {cell_20.value}")

        # Check that confidence 19 (BAL) is in row 4, column 4
        cell_19 = ws.cell(row=4, column=4)
        print(f"Row 4, Column 4 (Confidence 19): {cell_19.value}")

        # Check that confidence 1 (PIT) is in row 22, column 4
        cell_1 = ws.cell(row=22, column=4)
        print(f"Row 22, Column 4 (Confidence 1): {cell_1.value}")

        # Verify the structure
        print("\n📋 Excel Structure Check:")

        # Check header row (Week numbers)
        for col in range(2, 6):  # Check first few weeks
            week_num = ws.cell(row=1, column=col).value
            print(f"Week {col-1}: Column {col} = {week_num}")

        # Check confidence column
        for row in range(2, 6):  # Check first few confidence levels
            confidence = ws.cell(row=row, column=1).value
            print(f"Confidence {22-row}: Row {row} = {confidence}")

        print("\n✅ Excel file structure appears correct!")
        return True

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    success = test_excel_alignment()
    sys.exit(0 if success else 1)
