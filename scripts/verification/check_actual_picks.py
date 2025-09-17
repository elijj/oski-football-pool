#!/usr/bin/env python3
"""
Check what picks are actually in the Excel file.
"""

import os

from openpyxl import load_workbook


def check_actual_picks():
    """Check what picks are actually in the Excel file."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print(f"📊 Actual picks in Pool Week 1: {excel_file}")
        print("=" * 50)

        # Check what's actually in the Excel file
        print("Row | Col | Team | Confidence")
        print("-" * 30)

        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value  # Column B (Week 1)
            confidence = ws.cell(row=row, column=1).value  # Column A (confidence)
            print(f"{row:3d} | {2:3d} | {str(team):4s} | {confidence}")

        print("=" * 50)

        # Check if there are any CFB games mixed in
        print("\n🔍 Looking for CFB games in the schedule:")
        print("Games from the prompt that might be CFB:")
        cfb_candidates = [
            "CAL@SDSU",
            "STAN@VA",
            "UW@WSU",
            "FLA@Mia,F",
            "CAL@STAN",
            "LOU@SMU",
            "USC@ORE",
            "UT@FLA",
            "KSU@UTAH",
            "UW@UCLA",
        ]

        for game in cfb_candidates:
            print(f"  - {game}")

        print("\n💡 Note: The current picks are all NFL teams.")
        print("   CFB games would need to be added to the picks file.")

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")


if __name__ == "__main__":
    check_actual_picks()
