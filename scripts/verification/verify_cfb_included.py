#!/usr/bin/env python3
"""
Verify Pool Week 1 picks now include the missing CFB teams.
"""

import os

from openpyxl import load_workbook


def verify_cfb_included():
    """Verify that CFB teams are now included in Pool Week 1 picks."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🏈 Pool Week 1 Picks with CFB Teams Included")
        print("=" * 60)
        print("Conf | Team | Row | Game Match | League")
        print("-" * 50)

        # Expected picks with CFB teams included
        expected_picks = [
            ("KC", 20, "KC@NYG", "NFL"),
            ("BAL", 19, "DET@BALT", "NFL"),
            ("LAR", 18, "LAR@PHIL", "NFL"),
            ("DAL", 17, "DAL@CHI", "NFL"),
            ("GB", 16, "GB@CLEV", "NFL"),
            ("PHI", 15, "PHIL@DAL", "NFL"),
            ("SF", 14, "ARIZ@SF", "NFL"),
            ("BUF", 13, "MIA@BUFF", "NFL"),
            ("MIA", 12, "MIA@BUFF", "NFL"),
            ("DET", 11, "DET@BALT", "NFL"),
            ("NO", 10, "NO@SEA", "NFL"),
            ("TB", 9, "NYJ@TB", "NFL"),
            ("ATL", 8, "ATL@CAR", "NFL"),
            ("CAR", 7, "ATL@CAR", "NFL"),
            ("ARI", 6, "ARIZ@SF", "NFL"),
            ("UW", 5, "UW@WSU", "CFB"),
            ("CAL", 4, "CAL@SDSU", "CFB"),
            ("STAN", 3, "STAN@VA", "CFB"),
            ("FLA", 2, "FLA@Mia,F", "CFB"),
            ("SDSU", 1, "CAL@SDSU", "CFB"),
        ]

        all_correct = True

        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value

            # Find expected pick
            expected_team = None
            expected_game = "N/A"
            expected_league = "N/A"

            for exp_team, exp_conf, exp_game, exp_league in expected_picks:
                if exp_conf == confidence:
                    expected_team = exp_team
                    expected_game = exp_game
                    expected_league = exp_league
                    break

            status = "✅" if team == expected_team else "❌"
            if team != expected_team:
                all_correct = False

            print(
                f"{confidence:4d} | {str(team):4s} | {row:3d} | {expected_game:12s} | {expected_league:3s} | {status}"
            )

        print("=" * 60)

        if all_correct:
            print("🎉 POOL WEEK 1 NOW INCLUDES CFB TEAMS!")
            print("✅ Missing CFB games now included:")
            print("   - CAL@SDSU (CAL, SDSU)")
            print("   - STAN@VA (STAN)")
            print("   - UW@WSU (UW)")
            print("   - FLA@Mia,F (FLA)")
            print("✅ Mix of NFL (high confidence) and CFB (lower confidence)")
            print("✅ All 20 picks correctly aligned")
        else:
            print("❌ Some picks still need adjustment")

        return all_correct

    except Exception as e:
        print(f"❌ Error: {e}")
        return False


if __name__ == "__main__":
    success = verify_cfb_included()
    if success:
        print("\n🏆 Pool Week 1 with CFB teams is ready!")
    else:
        print("\n💥 Pool Week 1 still needs CFB teams!")
