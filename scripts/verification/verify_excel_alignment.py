#!/usr/bin/env python3
"""
Final verification that Excel alignment is correct.
"""

import os

from openpyxl import load_workbook


def verify_excel_alignment():
    """Verify that Excel file has correct pick alignment."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print(f"✅ Verifying Excel file: {excel_file}")
        print("=" * 60)

        # Expected picks from week_3_manual.json
        expected_picks = [
            ("KC", 20, 3, 4),  # KC, confidence 20, row 3, column 4
            ("BAL", 19, 4, 4),  # BAL, confidence 19, row 4, column 4
            ("LAR", 18, 5, 4),  # LAR, confidence 18, row 5, column 4
            ("DAL", 17, 6, 4),  # DAL, confidence 17, row 6, column 4
            ("GB", 16, 7, 4),  # GB, confidence 16, row 7, column 4
            ("PHI", 15, 8, 4),  # PHI, confidence 15, row 8, column 4
            ("SF", 14, 9, 4),  # SF, confidence 14, row 9, column 4
            ("BUF", 13, 10, 4),  # BUF, confidence 13, row 10, column 4
            ("MIA", 12, 11, 4),  # MIA, confidence 12, row 11, column 4
            ("DET", 11, 12, 4),  # DET, confidence 11, row 12, column 4
            ("ND", 10, 13, 4),  # ND, confidence 10, row 13, column 4
            ("TB", 9, 14, 4),  # TB, confidence 9, row 14, column 4
            ("ATL", 8, 15, 4),  # ATL, confidence 8, row 15, column 4
            ("CAR", 7, 16, 4),  # CAR, confidence 7, row 16, column 4
            ("ARI", 6, 17, 4),  # ARI, confidence 6, row 17, column 4
            ("SEA", 5, 18, 4),  # SEA, confidence 5, row 18, column 4
            ("LAC", 4, 19, 4),  # LAC, confidence 4, row 19, column 4
            ("LV", 3, 20, 4),  # LV, confidence 3, row 20, column 4
            ("DEN", 2, 21, 4),  # DEN, confidence 2, row 21, column 4
            ("PIT", 1, 22, 4),  # PIT, confidence 1, row 22, column 4
        ]

        print("🎯 Checking pick alignment:")
        print("Team | Conf | Row | Col | Expected | Actual | Status")
        print("-" * 60)

        all_correct = True

        for team, confidence, expected_row, expected_col in expected_picks:
            actual_value = ws.cell(row=expected_row, column=expected_col).value
            status = "✅" if actual_value == team else "❌"

            if actual_value != team:
                all_correct = False

            print(
                f"{team:4s} | {confidence:4d} | {expected_row:3d} | {expected_col:3d} | {team:8s} | {str(actual_value):6s} | {status}"
            )

        print("=" * 60)

        if all_correct:
            print("🎉 ALL PICKS CORRECTLY ALIGNED!")
            print("✅ Confidence 20 (KC) is in Row 3, Column 4")
            print("✅ Confidence 19 (BAL) is in Row 4, Column 4")
            print("✅ Confidence 1 (PIT) is in Row 22, Column 4")
            print("✅ All picks align with their confidence values")
        else:
            print("❌ SOME PICKS ARE MISALIGNED!")
            print("Please check the Excel file for incorrect placements")

        return all_correct

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    success = verify_excel_alignment()
    if success:
        print("\n🏆 Excel alignment verification PASSED!")
    else:
        print("\n💥 Excel alignment verification FAILED!")
