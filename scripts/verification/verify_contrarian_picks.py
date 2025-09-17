#!/usr/bin/env python3
"""
Verify contrarian picks are correctly placed in the Excel file.
"""

import os

from openpyxl import load_workbook


def verify_contrarian_picks():
    """Verify contrarian picks are correctly placed in the Excel file."""
    excel_file = "Dawgpac25_2024-09-17.xlsx"

    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return False

    try:
        wb = load_workbook(excel_file)
        ws = wb.active

        print("🎯 CONTRARIAN PICKS VERIFICATION")
        print("=" * 60)
        print(f"📁 File: {excel_file}")
        print("📅 Date: 2024-09-17 (Pool Week 1)")
        print("🎯 Strategy: Contrarian Analysis for Optimal Performance")
        print()

        # Expected contrarian picks
        expected_picks = [
            ("KC", 20, "HIGH CONFIDENCE - SAFETY FIRST"),
            ("BAL", 19, "HIGH CONFIDENCE - SAFETY FIRST"),
            ("LAR", 18, "HIGH CONFIDENCE - SAFETY FIRST"),
            ("DAL", 17, "HIGH CONFIDENCE - SAFETY FIRST"),
            ("GB", 16, "HIGH CONFIDENCE - SAFETY FIRST"),
            ("SF", 15, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("MIA", 14, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("BUF", 13, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("DET", 12, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("NO", 11, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("TB", 10, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("ATL", 9, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("CAR", 8, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("ARI", 7, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("SEA", 6, "MEDIUM CONFIDENCE - VALUE PLAYS"),
            ("LAC", 5, "LOW CONFIDENCE - UPSIDE PLAYS"),
            ("LV", 4, "LOW CONFIDENCE - UPSIDE PLAYS"),
            ("DEN", 3, "LOW CONFIDENCE - UPSIDE PLAYS"),
            ("WASH", 2, "LOW CONFIDENCE - UPSIDE PLAYS"),
            ("PITT", 1, "LOW CONFIDENCE - UPSIDE PLAYS"),
        ]

        print("📊 CONTRARIAN PICKS VERIFICATION:")
        print("Conf | Team | Row | Strategy | Expected | Actual | Status")
        print("-" * 70)

        all_correct = True

        for row in range(3, 23):  # Rows 3-22 (confidence 20-1)
            team = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=1).value

            # Find expected pick
            expected_team = None
            expected_strategy = "N/A"

            for exp_team, exp_conf, exp_strategy in expected_picks:
                if exp_conf == confidence:
                    expected_team = exp_team
                    expected_strategy = exp_strategy
                    break

            status = "✅" if team == expected_team else "❌"
            if team != expected_team:
                all_correct = False

            print(
                f"{confidence:4d} | {str(team):4s} | {row:3d} | {expected_strategy:25s} | {expected_team:8s} | {str(team):6s} | {status}"
            )

        print("=" * 70)

        if all_correct:
            print("🎉 CONTRARIAN PICKS CORRECTLY PLACED!")
            print("✅ High Confidence (20-16): SAFETY FIRST - 5 picks")
            print("✅ Medium Confidence (15-6): VALUE PLAYS - 10 picks")
            print("✅ Low Confidence (5-1): UPSIDE PLAYS - 5 picks")
            print("✅ Contrarian strategy implemented successfully")
            print("✅ Ready for Pool Week 1 submission")
        else:
            print("❌ SOME PICKS ARE MISALIGNED!")
            print("Please check the Excel file for incorrect placements")

        return all_correct

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False


if __name__ == "__main__":
    success = verify_contrarian_picks()
    if success:
        print("\n🏆 CONTRARIAN PICKS VERIFICATION PASSED!")
    else:
        print("\n💥 CONTRARIAN PICKS VERIFICATION FAILED!")
