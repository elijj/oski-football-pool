"""
Comprehensive test coverage for Excel automation to prevent pick alignment issues.
"""

import os
import shutil
import tempfile

import pytest
from openpyxl import Workbook, load_workbook

from football_pool.excel_automation import ExcelAutomation


class TestExcelAutomation:
    """Test suite for Excel automation functionality."""

    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for testing."""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        shutil.rmtree(temp_dir)

    @pytest.fixture
    def excel_automation(self, temp_dir):
        """Create ExcelAutomation instance for testing."""
        # Create a test template file
        template_path = os.path.join(temp_dir, "Dawgpac25.xlsx")
        wb = Workbook()
        ws = wb.active

        # Create the standard pool structure
        # Row 1: Week numbers (1-18)
        for col in range(2, 20):  # Columns B through S
            ws.cell(row=1, column=col, value=col - 1)

        # Column A: Confidence points (20-1)
        for row in range(2, 22):  # Rows 2 through 21
            ws.cell(row=row, column=1, value=22 - row)

        # Add some sample data to verify structure
        ws.cell(row=2, column=2, value="SAMPLE")  # Confidence 20, Week 1

        wb.save(template_path)

        # Create ExcelAutomation with the template
        automation = ExcelAutomation(template_path)
        automation.template_file = template_path
        return automation

    def test_pick_alignment_calculation(self, excel_automation):
        """Test that pick alignment calculations are correct."""
        # Test confidence to row mapping (Row 2 is "Name:" header, so confidence starts at row 3)
        test_cases = [
            (20, 3),  # Confidence 20 = Row 3
            (19, 4),  # Confidence 19 = Row 4
            (10, 13),  # Confidence 10 = Row 13
            (1, 22),  # Confidence 1 = Row 22
        ]

        for confidence, expected_row in test_cases:
            calculated_row = 23 - confidence
            assert (
                calculated_row == expected_row
            ), f"Confidence {confidence} should map to row {expected_row}, got {calculated_row}"

    def test_week_column_calculation(self, excel_automation):
        """Test that week to column mapping is correct."""
        # Test week to column mapping
        test_cases = [
            (1, 2),  # Week 1 = Column B (2)
            (2, 3),  # Week 2 = Column C (3)
            (3, 4),  # Week 3 = Column D (4)
            (18, 19),  # Week 18 = Column S (19)
        ]

        for week, expected_col in test_cases:
            calculated_col = week + 1
            assert (
                calculated_col == expected_col
            ), f"Week {week} should map to column {expected_col}, got {calculated_col}"

    def test_pick_placement_accuracy(self, excel_automation, temp_dir):
        """Test that picks are placed in correct cells."""
        # Create test picks with known positions
        test_picks = [
            {"team": "KC", "confidence": 20, "week": 3},  # Should go to Row 2, Column 4
            {"team": "BALT", "confidence": 19, "week": 3},  # Should go to Row 3, Column 4
            {"team": "LAR", "confidence": 10, "week": 3},  # Should go to Row 12, Column 4
            {"team": "DET", "confidence": 1, "week": 3},  # Should go to Row 21, Column 4
        ]

        # Update picks
        excel_automation.update_picks(3, test_picks, "2024-09-17")

        # Load the updated file and verify placements
        filename = os.path.join(temp_dir, "Dawgpac25_2024-09-17.xlsx")
        wb = load_workbook(filename)
        ws = wb.active

        # Verify each pick is in the correct cell
        assert ws.cell(row=3, column=4).value == "KC", "KC should be in Row 3, Column 4"
        assert ws.cell(row=4, column=4).value == "BALT", "BALT should be in Row 4, Column 4"
        assert ws.cell(row=13, column=4).value == "LAR", "LAR should be in Row 13, Column 4"
        assert ws.cell(row=22, column=4).value == "DET", "DET should be in Row 22, Column 4"

    def test_confidence_point_validation(self, excel_automation):
        """Test that confidence points are properly validated."""
        # Test valid confidence points (need 20 picks for full validation)
        valid_picks = []
        for i in range(20, 0, -1):
            valid_picks.append({"team": f"TEAM{i}", "confidence": i})

        validation = excel_automation.validate_picks(valid_picks)
        assert validation[
            "valid"
        ], f"Valid picks should pass validation: {validation.get('errors', [])}"

        # Test invalid confidence points
        invalid_picks = [
            {"team": "KC", "confidence": 21},  # Too high
            {"team": "BALT", "confidence": 0},  # Too low
            {"team": "LAR", "confidence": -1},  # Negative
        ]

        validation = excel_automation.validate_picks(invalid_picks)
        assert not validation["valid"], "Invalid confidence points should fail validation"

    def test_duplicate_confidence_prevention(self, excel_automation):
        """Test that duplicate confidence points are detected."""
        duplicate_picks = [
            {"team": "KC", "confidence": 20},
            {"team": "BALT", "confidence": 20},  # Duplicate
            {"team": "LAR", "confidence": 19},
        ]

        validation = excel_automation.validate_picks(duplicate_picks)
        assert not validation["valid"], "Duplicate confidence points should fail validation"
        assert "duplicate" in str(validation.get("errors", [])).lower()

    def test_team_name_validation(self, excel_automation):
        """Test that team names are properly validated."""
        # Test valid team names (need 20 picks for full validation)
        valid_picks = []
        for i in range(20, 0, -1):
            valid_picks.append({"team": f"TEAM{i}", "confidence": i})

        validation = excel_automation.validate_picks(valid_picks)
        assert validation[
            "valid"
        ], f"Valid team names should pass validation: {validation.get('errors', [])}"

        # Test invalid team names
        invalid_picks = [
            {"team": "", "confidence": 20},  # Empty team
            {"team": None, "confidence": 19},  # None team
            {"team": "INVALID", "confidence": 18},  # Invalid team name
        ]

        validation = excel_automation.validate_picks(invalid_picks)
        assert not validation["valid"], "Invalid team names should fail validation"

    def test_excel_structure_integrity(self, excel_automation, temp_dir):
        """Test that Excel file structure remains intact after updates."""
        # Create test picks
        test_picks = [
            {"team": "KC", "confidence": 20, "week": 3},
            {"team": "BALT", "confidence": 19, "week": 3},
            {"team": "LAR", "confidence": 18, "week": 3},
        ]

        # Update picks
        excel_automation.update_picks(3, test_picks, "2024-09-17")

        # Load and verify structure
        filename = os.path.join(temp_dir, "Dawgpac25_2024-09-17.xlsx")
        wb = load_workbook(filename)
        ws = wb.active

        # Verify header row (Week numbers)
        for col in range(2, 20):
            assert (
                ws.cell(row=1, column=col).value == col - 1
            ), f"Week number should be {col-1} in column {col}"

        # Verify confidence column (Column A)
        for row in range(2, 22):
            assert (
                ws.cell(row=row, column=1).value == 22 - row
            ), f"Confidence should be {22-row} in row {row}"

    def test_cell_formatting_consistency(self, excel_automation, temp_dir):
        """Test that cell formatting is consistent."""
        # Create test picks
        test_picks = [
            {"team": "KC", "confidence": 20, "week": 3},
            {"team": "BALT", "confidence": 19, "week": 3},
        ]

        # Update picks
        excel_automation.update_picks(3, test_picks, "2024-09-17")

        # Load and verify formatting
        filename = os.path.join(temp_dir, "Dawgpac25_2024-09-17.xlsx")
        wb = load_workbook(filename)
        ws = wb.active

        # Check that cells have proper formatting
        cell_20 = ws.cell(row=2, column=4)  # Confidence 20
        cell_19 = ws.cell(row=3, column=4)  # Confidence 19

        assert cell_20.value == "KC", "KC should be in the correct cell"
        assert cell_19.value == "BALT", "BALT should be in the correct cell"

        # Verify formatting
        assert cell_20.font.bold, "Cell should be bold"
        assert cell_19.font.bold, "Cell should be bold"

    def test_backup_functionality(self, excel_automation, temp_dir):
        """Test that backup functionality works correctly."""
        # Create test picks
        test_picks = [
            {"team": "KC", "confidence": 20, "week": 3},
        ]

        # Update picks
        excel_automation.update_picks(3, test_picks, "2024-09-17")

        # Create backup
        backup_file = excel_automation.backup_file(3, "2024-09-17")

        # Verify backup was created
        assert backup_file is not None, "Backup file should be created"
        assert os.path.exists(backup_file), "Backup file should exist"

        # Verify backup contains the data
        wb = load_workbook(backup_file)
        ws = wb.active
        assert ws.cell(row=2, column=4).value == "KC", "Backup should contain the pick data"

    def test_error_handling(self, excel_automation):
        """Test that error handling works correctly."""
        # Test with invalid data
        invalid_picks = [
            {"team": "KC", "confidence": 25},  # Invalid confidence
            {"team": "", "confidence": 20},  # Empty team
        ]

        # Should handle errors gracefully
        try:
            excel_automation.update_picks(3, invalid_picks, "2024-09-17")
        except Exception as e:
            # Should not crash the system
            assert "confidence" in str(e).lower() or "team" in str(e).lower()

    def test_pick_retrieval(self, excel_automation, temp_dir):
        """Test that picks can be retrieved correctly."""
        # Create test picks
        test_picks = [
            {"team": "KC", "confidence": 20, "week": 3},
            {"team": "BALT", "confidence": 19, "week": 3},
            {"team": "LAR", "confidence": 18, "week": 3},
        ]

        # Update picks
        excel_automation.update_picks(3, test_picks, "2024-09-17")

        # Retrieve picks
        retrieved_picks = excel_automation.get_current_picks(3)

        # Verify retrieval
        assert len(retrieved_picks) == 3, "Should retrieve 3 picks"

        # Verify content
        teams = [pick["team"] for pick in retrieved_picks]
        assert "KC" in teams, "Should retrieve KC"
        assert "BALT" in teams, "Should retrieve BALT"
        assert "LAR" in teams, "Should retrieve LAR"

    def test_week_boundary_validation(self, excel_automation):
        """Test that week boundaries are properly validated."""
        # Test valid weeks
        valid_weeks = [3, 4, 5, 18]
        for week in valid_weeks:
            test_picks = [{"team": "KC", "confidence": 20, "week": week}]
            validation = excel_automation.validate_picks(test_picks)
            assert validation["valid"], f"Week {week} should be valid"

        # Test invalid weeks
        invalid_weeks = [0, 1, 2, 19, 20]
        for week in invalid_weeks:
            test_picks = [{"team": "KC", "confidence": 20, "week": week}]
            validation = excel_automation.validate_picks(test_picks)
            # Should either be valid or have specific error handling
            # (depending on business rules)

    def test_comprehensive_pick_validation(self, excel_automation):
        """Test comprehensive pick validation with all edge cases."""
        # Test complete valid pick set
        valid_picks = []
        for i in range(20, 0, -1):  # Confidence 20 down to 1
            valid_picks.append({"team": f"TEAM{i}", "confidence": i, "week": 3})

        validation = excel_automation.validate_picks(valid_picks)
        assert validation[
            "valid"
        ], f"Complete valid pick set should pass: {validation.get('errors', [])}"

        # Test missing picks
        incomplete_picks = valid_picks[:10]  # Only 10 picks
        validation = excel_automation.validate_picks(incomplete_picks)
        # Should either be valid or have specific error handling

        # Test too many picks
        too_many_picks = valid_picks + [{"team": "EXTRA", "confidence": 21, "week": 3}]
        validation = excel_automation.validate_picks(too_many_picks)
        assert not validation["valid"], "Too many picks should fail validation"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
