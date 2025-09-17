"""
Excel Automation for Football Pool Submission

Handles the automated creation and updating of Excel files for weekly picks submission
following the exact format required by the pool organizer.
"""

import logging
import os
import shutil
import csv
import json
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

logger = logging.getLogger(__name__)


class ExcelAutomation:
    """Handles Excel file automation for pool submissions."""

    def __init__(self, template_path: str = "data/excel/Dawgpac25.xlsx"):
        """Initialize with template file path."""
        self.template_path = template_path
        self.output_dir = "data/excel"
        self.workbook = None
        self.worksheet = None

    def load_template(self) -> bool:
        """Load the Excel template file."""
        try:
            if not os.path.exists(self.template_path):
                logger.error(f"Template file not found: {self.template_path}")
                return False

            self.workbook = load_workbook(self.template_path)
            self.worksheet = self.workbook.active
            logger.info(f"Template loaded successfully: {self.template_path}")
            return True

        except Exception as e:
            logger.error(f"Error loading template: {e}")
            return False

    def create_weekly_file(
        self, week: int, date: str = None, participant_name: str = "Dawgpac"
    ) -> str:
        """Create a new weekly file based on template with date suffix."""
        try:
            # Load template
            if not self.load_template():
                return None

            # Create filename with date suffix
            if date:
                filename = f"Dawgpac25_{date}.xlsx"
            else:
                # Default to week number if no date provided
                filename = f"Dawgpac25_Week{week}.xlsx"

            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)

            # Full path for output
            output_path = os.path.join(self.output_dir, filename)

            # Save as new file (preserving original template)
            self.workbook.save(output_path)
            logger.info(f"Created weekly file: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error creating weekly file: {e}")
            return None

    def update_picks(
        self,
        week: int,
        picks: list[dict[str, Any]],
        date: str = None,
        participant_name: str = "Dawgpac",
    ) -> bool:
        """Update picks for a specific week in the Excel file."""
        try:
            # Determine filename
            if date:
                filename = f"Dawgpac25_{date}.xlsx"
            else:
                filename = f"Dawgpac25_Week{week}.xlsx"

            # Full path for file
            file_path = os.path.join(self.output_dir, filename)

            if not os.path.exists(file_path):
                # Create new file if it doesn't exist
                file_path = self.create_weekly_file(week, date, participant_name)
                if not file_path:
                    return False

            # Load workbook
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Update picks in the Excel file
            # The structure is: Row 1 = Week numbers (1-18), Column A = Confidence points (20-1)
            # We need to find the right cell for each pick

            for pick in picks:
                team = pick.get("team", "")
                confidence = pick.get("confidence", 0)
                week_num = pick.get("week", week)

                if team and confidence:
                    # Find the row for this confidence level
                    # Confidence 20 = row 3, Confidence 1 = row 22
                    # Row 2 is "Name:" header, so confidence starts at row 3
                    row = 23 - confidence

                    # Find the column for this week
                    # Week 1 = column 2 (B), Week 2 = column 3 (C), etc.
                    col = week_num + 1

                    # Update the cell
                    cell = worksheet.cell(row=row, column=col)
                    cell.value = team

                    # Style the cell to match row 2 formatting
                    if row == 2:
                        # For row 2, match the formatting of other cells in that row
                        cell.font = Font(bold=True, size=11)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin"),
                        )
                    else:
                        # For other rows, use standard formatting
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center")

            # Save the updated file
            workbook.save(file_path)
            logger.info(f"Updated picks for Week {week} in {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error updating picks: {e}")
            return False

    def get_current_picks(self, week: int, date: str = None) -> list[dict[str, Any]]:
        """Get current picks from the Excel file for a specific week."""
        try:
            # Determine filename
            if date:
                filename = f"Dawgpac25_{date}.xlsx"
            else:
                filename = f"Dawgpac25_Week{week}.xlsx"

            # Full path for file
            file_path = os.path.join(self.output_dir, filename)

            if not os.path.exists(file_path):
                return []

            # Load workbook
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            picks = []

            # Read picks from the Excel file
            for row in range(2, 22):  # Rows 2-21 (confidence 20-1)
                confidence = 22 - row  # Calculate confidence from row
                cell = worksheet.cell(row=row, column=week + 1)
                team = cell.value

                if team and str(team).strip():
                    picks.append(
                        {"team": str(team).strip(), "confidence": confidence, "week": week}
                    )

            return picks

        except Exception as e:
            logger.error(f"Error reading picks: {e}")
            return []

    def validate_picks(self, picks: list[dict[str, Any]]) -> dict[str, Any]:
        """Validate picks according to pool rules."""
        validation_result = {"valid": True, "errors": [], "warnings": []}

        # Check if we have exactly 20 picks
        if len(picks) != 20:
            validation_result["valid"] = False
            validation_result["errors"].append(f"Expected 20 picks, got {len(picks)}")

        # Check confidence points (1-20, no duplicates)
        confidences = [pick.get("confidence", 0) for pick in picks]
        if len(set(confidences)) != len(confidences):
            validation_result["valid"] = False
            validation_result["errors"].append("Duplicate confidence points found")

        if min(confidences) < 1 or max(confidences) > 20:
            validation_result["valid"] = False
            validation_result["errors"].append("Confidence points must be between 1 and 20")

        # Check for empty teams
        empty_teams = [i for i, pick in enumerate(picks) if not pick.get("team", "").strip()]
        if empty_teams:
            validation_result["valid"] = False
            validation_result["errors"].append(f"Empty team names at positions: {empty_teams}")

        return validation_result

    def create_submission_summary(self, week: int, date: str = None) -> str:
        """Create a summary of picks for submission."""
        try:
            picks = self.get_current_picks(week)
            if not picks:
                return "No picks found for this week."

            # Sort by confidence (highest first)
            picks.sort(key=lambda x: x["confidence"], reverse=True)

            summary = f"Week {week} Picks Summary:\n"
            summary += "=" * 30 + "\n"

            for pick in picks:
                summary += f"{pick['confidence']:2d}: {pick['team']}\n"

            return summary

        except Exception as e:
            logger.error(f"Error creating submission summary: {e}")
            return f"Error creating summary: {e}"

    def backup_file(self, week: int, date: str = None) -> str:
        """Create a backup of the current file."""
        try:
            filename = f"Dawgpac25_Week{week}.xlsx"
            file_path = os.path.join(self.output_dir, filename)

            if not os.path.exists(file_path):
                return None

            # Create backup with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"Dawgpac25_Week{week}_backup_{timestamp}.xlsx"
            backup_path = os.path.join(self.output_dir, backup_filename)

            shutil.copy2(file_path, backup_path)
            logger.info(f"Created backup: {backup_path}")
            return backup_path

        except Exception as e:
            logger.error(f"Error creating backup: {e}")
            return None

    def get_team_abbreviations(self) -> dict[str, str]:
        """Get standard team abbreviations for the pool."""
        return {
            # NFL Teams
            "Kansas City Chiefs": "KC",
            "New York Giants": "NYG",
            "Buffalo Bills": "BUF",
            "Miami Dolphins": "MIA",
            "Los Angeles Rams": "LAR",
            "San Francisco 49ers": "SF",
            "Dallas Cowboys": "DAL",
            "Philadelphia Eagles": "PHI",
            "Green Bay Packers": "GB",
            "Chicago Bears": "CHI",
            "Detroit Lions": "DET",
            "Minnesota Vikings": "MIN",
            "New Orleans Saints": "NO",
            "Tampa Bay Buccaneers": "TB",
            "Atlanta Falcons": "ATL",
            "Carolina Panthers": "CAR",
            "Arizona Cardinals": "ARI",
            "Seattle Seahawks": "SEA",
            "Los Angeles Chargers": "LAC",
            "Las Vegas Raiders": "LV",
            "Denver Broncos": "DEN",
            "Pittsburgh Steelers": "PIT",
            "PITT": "PIT",  # Handle PITT -> PIT conversion
            "Baltimore Ravens": "BAL",
            "Cleveland Browns": "CLE",
            "Cincinnati Bengals": "CIN",
            "Houston Texans": "HOU",
            "Indianapolis Colts": "IND",
            "Tennessee Titans": "TEN",
            "Jacksonville Jaguars": "JAX",
            "New England Patriots": "NE",
            "New York Jets": "NYJ",
            "Washington Commanders": "WAS",
            "WASH": "WAS",  # Handle WASH -> WAS conversion
            # CFB Teams (common abbreviations)
            "Alabama": "ALA",
            "Georgia": "UGA",
            "Ohio State": "OSU",
            "Michigan": "MICH",
            "Clemson": "CLEM",
            "Notre Dame": "ND",
            "Oklahoma": "OU",
            "Texas": "TEX",
            "USC": "USC",
            "LSU": "LSU",
            "Florida": "UF",
            "Auburn": "AUB",
            "Tennessee": "TENN",
            "Kentucky": "UK",
            "South Carolina": "SC",
            "Missouri": "MIZ",
            "Arkansas": "ARK",
            "Mississippi State": "MSST",
            "Ole Miss": "MISS",
            "Vanderbilt": "VANDY",
            # Additional CFB teams
            "Louisville": "LOU",
            "Stanford": "STAN",
            "Penn State": "PSU",
            "UCLA": "UCLA",
            "Florida State": "FSU",
            "Nebraska": "NEB",
            "Indiana": "UIND",
            "California": "CAL",
            "North Carolina": "NC",
            "Alabama": "ALA",
            "Washington": "UW",
            "Washington State": "WSU",
            "UW": "UW",  # Handle UW -> UW conversion
            "WSU": "WSU",  # Handle WSU -> WSU conversion
        }

    def convert_team_names(self, picks: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Convert team names to standard abbreviations."""
        abbreviations = self.get_team_abbreviations()

        converted_picks = []
        for pick in picks:
            team = pick.get("team", "")
            original_team = team

            # First, check if it's already an abbreviation (exact match)
            if team.upper() in [abbrev.upper() for abbrev in abbreviations.values()]:
                # Already an abbreviation, keep as is
                converted_picks.append(pick)
                continue

            # Try to find abbreviation with more precise matching
            found_match = False
            for full_name, abbrev in abbreviations.items():
                # More precise matching to avoid conflicts like NO -> ND
                if team.upper() == abbrev.upper() or (
                    len(team) > 2
                    and team.lower() in full_name.lower()
                    and abs(len(team) - len(abbrev)) <= 2
                ):
                    pick["team"] = abbrev
                    found_match = True
                    break

            if not found_match:
                # Keep original if no match found
                logger.warning(f"No abbreviation found for team: {team}")

            converted_picks.append(pick)

        return converted_picks

    def load_contrarian_analysis(self, analysis_file: str) -> list[dict[str, Any]]:
        """Load picks from contrarian analysis JSON file."""
        try:
            import json

            with open(analysis_file) as f:
                data = json.load(f)

            picks = []
            if "optimal_picks" in data:
                for pick in data["optimal_picks"]:
                    picks.append(
                        {
                            "game": pick.get("game", ""),
                            "team": pick.get("team", ""),
                            "confidence": pick.get("confidence", 0),
                            "reasoning": pick.get("reasoning", ""),
                            "contrarian_edge": pick.get("contrarian_edge", ""),
                            "value_play": pick.get("value_play", ""),
                            "risk_assessment": pick.get("risk_assessment", ""),
                        }
                    )

            logger.info(f"Loaded {len(picks)} picks from contrarian analysis")
            return picks

        except Exception as e:
            logger.error(f"Error loading contrarian analysis: {e}")
            return []

    def update_picks_from_analysis(self, analysis_file: str, week: int = 1, date: str = None) -> bool:
        """Update Excel file with picks from analysis file."""
        try:
            # Load picks from analysis
            picks = self.load_contrarian_analysis(analysis_file)

            if not picks:
                print("No picks found in analysis file")
                return False

            # Create weekly file first
            if not date:
                date = "2025-09-17"  # Default date
            weekly_file = self.create_weekly_file(week, date)

            if not weekly_file:
                print("Failed to create weekly file")
                return False

            # Load the created file
            if not self.load_template():
                print("Failed to load template")
                return False

            # Update picks
            return self.update_picks(picks, week, date)

        except Exception as e:
            print(f"Error updating picks from analysis: {e}")
            return False

    def create_comprehensive_csv(
        self, week: int, date: str, analysis_file: str = None
    ) -> str:
        """Create a comprehensive CSV with all contrarian analysis metadata."""
        try:
            # Load contrarian analysis if provided
            picks_data = []
            if analysis_file and os.path.exists(analysis_file):
                picks_data = self.load_contrarian_analysis(analysis_file)
            else:
                # Fallback to basic picks if no analysis provided
                logger.warning("No contrarian analysis provided, using basic picks")
                return None

            if not picks_data:
                logger.error("No picks data available for CSV generation")
                return None

            # Create comprehensive CSV filename
            csv_filename = f"Week_{week}_Picks_{date}.csv"
            csv_path = os.path.join(self.output_dir, csv_filename)

            # Ensure output directory exists
            os.makedirs(self.output_dir, exist_ok=True)

            # Define comprehensive CSV headers - put long text columns at the end
            headers = [
                "Game",
                "Team",
                "Confidence_Points",
                "Confidence_Percentage",
                "Strategy_Type",
                "Weather_Impact",
                "Injury_Impact",
                "Public_Betting_Percentage",
                "Sharp_Money_Indicator",
                "Home_Field_Advantage",
                "Matchup_Advantage",
                "Historical_Performance",
                "Recent_Form",
                "Key_Player_Status",
                "Weather_Conditions",
                "Injury_Report",
                "Line_Movement",
                "Public_Sentiment",
                "Expert_Consensus",
                "Value_Rating",
                "Upside_Potential",
                "Downside_Risk",
                "Risk_Assessment",
                "Reasoning",
                "Contrarian_Edge",
                "Value_Play"
            ]

            # Write comprehensive CSV
            with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()

                for pick in picks_data:
                    # Extract all available metadata
                    row = {
                        "Game": pick.get("game", ""),
                        "Team": pick.get("team", ""),
                        "Confidence_Points": pick.get("confidence", ""),
                        "Confidence_Percentage": f"{pick.get('confidence', 0) * 5:.1f}%",  # Convert to percentage
                        "Reasoning": pick.get("reasoning", ""),
                        "Contrarian_Edge": pick.get("contrarian_edge", ""),
                        "Value_Play": pick.get("value_play", ""),
                        "Risk_Assessment": pick.get("risk_assessment", ""),
                        "Strategy_Type": self._determine_strategy_type(pick),
                        "Weather_Impact": self._get_weather_impact(pick),
                        "Injury_Impact": self._get_injury_impact(pick),
                        "Public_Betting_Percentage": self._get_public_betting(pick),
                        "Sharp_Money_Indicator": self._get_sharp_money(pick),
                        "Home_Field_Advantage": self._get_home_field_advantage(pick),
                        "Matchup_Advantage": self._get_matchup_advantage(pick),
                        "Historical_Performance": self._get_historical_performance(pick),
                        "Recent_Form": self._get_recent_form(pick),
                        "Key_Player_Status": self._get_key_player_status(pick),
                        "Weather_Conditions": self._get_weather_conditions(pick),
                        "Injury_Report": self._get_injury_report(pick),
                        "Line_Movement": self._get_line_movement(pick),
                        "Public_Sentiment": self._get_public_sentiment(pick),
                        "Expert_Consensus": self._get_expert_consensus(pick),
                        "Value_Rating": self._get_value_rating(pick),
                        "Upside_Potential": self._get_upside_potential(pick),
                        "Downside_Risk": self._get_downside_risk(pick)
                    }
                    writer.writerow(row)

            logger.info(f"Comprehensive CSV created: {csv_path}")
            return csv_path

        except Exception as e:
            logger.error(f"Error creating comprehensive CSV: {e}")
            return None

    def _determine_strategy_type(self, pick: Dict) -> str:
        """Determine strategy type based on pick data."""
        confidence = pick.get("confidence", 0)
        risk = pick.get("risk_assessment", "").lower()

        if confidence >= 18:
            return "HIGH_CONFIDENCE_SAFETY"
        elif confidence >= 15:
            return "MEDIUM_CONFIDENCE_VALUE"
        elif "high" in risk:
            return "HIGH_RISK_UPSIDE"
        else:
            return "BALANCED_PLAY"

    def _get_weather_impact(self, pick: Dict) -> str:
        """Get weather impact for the pick."""
        # This would integrate with weather data provider
        return "MODERATE"  # Placeholder

    def _get_injury_impact(self, pick: Dict) -> str:
        """Get injury impact for the pick."""
        # This would integrate with injury data provider
        return "LOW"  # Placeholder

    def _get_public_betting(self, pick: Dict) -> str:
        """Get public betting percentage."""
        return "65%"  # Placeholder

    def _get_sharp_money(self, pick: Dict) -> str:
        """Get sharp money indicator."""
        return "YES" if pick.get("confidence", 0) >= 15 else "NO"

    def _get_home_field_advantage(self, pick: Dict) -> str:
        """Get home field advantage assessment."""
        game = pick.get("game", "")
        if "@" in game:
            home_team = game.split("@")[1]
            picked_team = pick.get("team", "")
            return "YES" if picked_team == home_team else "NO"
        return "UNKNOWN"

    def _get_matchup_advantage(self, pick: Dict) -> str:
        """Get matchup advantage assessment."""
        return "STRONG" if pick.get("confidence", 0) >= 15 else "MODERATE"

    def _get_historical_performance(self, pick: Dict) -> str:
        """Get historical performance data."""
        return "STRONG"  # Placeholder

    def _get_recent_form(self, pick: Dict) -> str:
        """Get recent form assessment."""
        return "GOOD"  # Placeholder

    def _get_key_player_status(self, pick: Dict) -> str:
        """Get key player status."""
        return "HEALTHY"  # Placeholder

    def _get_weather_conditions(self, pick: Dict) -> str:
        """Get weather conditions."""
        return "CLEAR"  # Placeholder

    def _get_injury_report(self, pick: Dict) -> str:
        """Get injury report."""
        return "CLEAN"  # Placeholder

    def _get_line_movement(self, pick: Dict) -> str:
        """Get line movement data."""
        return "STABLE"  # Placeholder

    def _get_public_sentiment(self, pick: Dict) -> str:
        """Get public sentiment."""
        return "MIXED"  # Placeholder

    def _get_expert_consensus(self, pick: Dict) -> str:
        """Get expert consensus."""
        return "FAVOR" if pick.get("confidence", 0) >= 15 else "MIXED"

    def _get_value_rating(self, pick: Dict) -> str:
        """Get value rating."""
        confidence = pick.get("confidence", 0)
        if confidence >= 18:
            return "EXCELLENT"
        elif confidence >= 15:
            return "GOOD"
        elif confidence >= 10:
            return "FAIR"
        else:
            return "POOR"

    def _get_upside_potential(self, pick: Dict) -> str:
        """Get upside potential."""
        return "HIGH" if pick.get("confidence", 0) >= 15 else "MODERATE"

    def _get_downside_risk(self, pick: Dict) -> str:
        """Get downside risk."""
        risk = pick.get("risk_assessment", "").lower()
        if "high" in risk:
            return "HIGH"
        elif "medium" in risk:
            return "MEDIUM"
        else:
            return "LOW"
